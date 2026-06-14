import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def hospitals_to_dataframe(scored_hospitals):
    rows = []

    for h in scored_hospitals:
        web_info = h.get("정리된웹병상정보", {})

        rows.append({
            "추천순위": h.get("추천순위"),
            "병원명": h.get("웹병원명"),
            "기관구분": h.get("기관구분"),
            "주소": h.get("주소"),
            "직선거리(km)": h.get("직선거리_km"),
            "실제이동거리(km)": h.get("이동거리_km"),
            "예상이동시간(분)": h.get("예상이동시간_분"),
            "추천점수": h.get("추천점수"),

            "응급실일반": web_info.get("응급실일반"),
            "응급실일반_혼잡도": web_info.get("응급실일반_혼잡도"),
            "응급실소아": web_info.get("응급실소아"),
            "응급실소아_혼잡도": web_info.get("응급실소아_혼잡도"),
            "음압격리": web_info.get("음압격리"),
            "일반격리": web_info.get("일반격리"),
            "코호트격리": web_info.get("코호트격리"),

            "중환자실_API": h.get("중환자실_API"),
            "수술실_API": h.get("수술실_API"),
            "CT가능": h.get("CT가능"),
            "MRI가능": h.get("MRI가능"),
            "인공호흡기": h.get("인공호흡기"),

            "API병원명": h.get("API병원명"),
            "매칭점수": h.get("매칭점수"),
            "추천사유": h.get("추천사유"),
            "업데이트시간": h.get("업데이트시간")
        })

    return pd.DataFrame(rows)


def save_recommendation_excel(scored_hospitals, filename=None):
    os.makedirs("output", exist_ok=True)

    if filename is None:
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"응급실_추천_결과_{now}.xlsx"

    file_path = os.path.join("output", filename)

    # 추천 결과를 DataFrame으로 변환
    df = hospitals_to_dataframe(scored_hospitals)

    # pandas로 엑셀 파일 생성
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="추천결과", index=False)

        workbook = writer.book
        ws = writer.sheets["추천결과"]

        # 기본 스타일 객체
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        top_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        warning_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        normal_font = Font(size=10)
        thin_border = Border(
            left=Side(style="thin", color="999999"),
            right=Side(style="thin", color="999999"),
            top=Side(style="thin", color="999999"),
            bottom=Side(style="thin", color="999999")
        )

        # 헤더 스타일 지정
        ws.row_dimensions[1].height = 34

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 본문 스타일 지정
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            rank = row[0].value
            score = row[7].value

            for cell in row:
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 주소와 추천사유는 내용이 길어 왼쪽 정렬
            row[3].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            row[22].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # TOP 5 병원 강조
            if isinstance(rank, int) and rank <= 5:
                for cell in row:
                    cell.fill = top_fill

            # 점수가 낮은 병원은 따로 표시
            if isinstance(score, (int, float)) and score < 30:
                for cell in row:
                    cell.fill = warning_fill

        # 열 너비 지정
        column_widths = {
            "A": 10, "B": 24, "C": 12, "D": 58,
            "E": 16, "F": 18, "G": 20, "H": 13,
            "I": 14, "J": 20, "K": 14, "L": 20,
            "M": 12, "N": 12, "O": 12,
            "P": 16, "Q": 14, "R": 10, "S": 10, "T": 14,
            "U": 38, "V": 12, "W": 90, "X": 22
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 행 높이 지정
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 42

        # 숫자 표시 형식
        for row_idx in range(2, ws.max_row + 1):
            ws[f"E{row_idx}"].number_format = "0.00"
            ws[f"F{row_idx}"].number_format = "0.00"
            ws[f"G{row_idx}"].number_format = "0.0"
            ws[f"H{row_idx}"].number_format = "0"
            ws[f"V{row_idx}"].number_format = "0.00"

        # 틀 고정과 필터
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # 요약 시트 생성
        ws_summary = workbook.create_sheet("요약")
        ws_summary["A1"] = "응급실 추천 결과 요약"
        ws_summary["A1"].font = Font(bold=True, size=15)

        ws_summary["A3"] = "전체 후보 병원 수"
        ws_summary["B3"] = len(df)

        ws_summary["A4"] = "최고 추천점수"
        ws_summary["B4"] = df["추천점수"].max()

        ws_summary["A5"] = "평균 추천점수"
        ws_summary["B5"] = round(df["추천점수"].mean(), 2)

        ws_summary["A6"] = "가장 가까운 병원"

        if df["실제이동거리(km)"].notna().any():
            nearest = df.sort_values("실제이동거리(km)").iloc[0]
            ws_summary["B6"] = nearest["병원명"]
            ws_summary["C6"] = nearest["실제이동거리(km)"]
        elif df["직선거리(km)"].notna().any():
            nearest = df.sort_values("직선거리(km)").iloc[0]
            ws_summary["B6"] = nearest["병원명"]
            ws_summary["C6"] = nearest["직선거리(km)"]

        ws_summary["A8"] = "TOP 5 병원"
        ws_summary["A8"].font = Font(bold=True, size=13)

        # append로 표 생성
        ws_summary.append([
            "순위",
            "병원명",
            "추천점수",
            "직선거리(km)",
            "실제이동거리(km)",
            "예상이동시간(분)",
            "추천사유"
        ])

        for _, row in df.head(5).iterrows():
            ws_summary.append([
                row["추천순위"],
                row["병원명"],
                row["추천점수"],
                row.get("직선거리(km)"),
                row.get("실제이동거리(km)"),
                row.get("예상이동시간(분)"),
                row["추천사유"]
            ])

        # 요약 시트 서식
        for row in ws_summary.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.font = normal_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 요약 시트 헤더 강조
        for cell in ws_summary[9]:
            cell.fill = header_fill
            cell.font = header_font

        # TOP5 영역 강조
        for row_idx in range(10, min(14, ws_summary.max_row) + 1):
            for cell in ws_summary[row_idx]:
                cell.fill = top_fill

        # 요약 시트 열 너비
        ws_summary.column_dimensions["A"].width = 12
        ws_summary.column_dimensions["B"].width = 28
        ws_summary.column_dimensions["C"].width = 15
        ws_summary.column_dimensions["D"].width = 18
        ws_summary.column_dimensions["E"].width = 20
        ws_summary.column_dimensions["F"].width = 20
        ws_summary.column_dimensions["G"].width = 90

        for row_idx in range(1, ws_summary.max_row + 1):
            ws_summary.row_dimensions[row_idx].height = 32

        ws_summary.freeze_panes = "A9"

    print(f"Excel 보고서 저장 완료: {file_path}")
    return file_path